"""Excel export for deep analysis results.

The parser pipeline intentionally stays unaware of workbook details.  This
module receives already collected pair states and turns them into two sheets:
an analyst-friendly matrix and a lossless detail sheet with every date pair.
"""
from __future__ import annotations

import json
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from loguru import logger


PRICED = "priced"
SOLD_OUT = "sold_out"
MIN_LOS = "min_los"

_PRICE_RE = re.compile(r"(\d[\d\s\u00a0\u202f]*)")
_DETAIL_COLUMNS = [
    "Объект",
    "ID",
    "Сайт",
    "Категория",
    "Ссылка",
    "Дата заезда",
    "Дата выезда",
    "Ночей",
    "Цена",
    "Статус",
    "Причина",
]


@dataclass(frozen=True)
class PairResult:
    checkin: date
    checkout: date
    nights: int
    status: str
    display_status: str
    price: Optional[float]
    reason: Optional[str]
    min_los_nights: Optional[int] = None


@dataclass(frozen=True)
class PropertyExportResult:
    id: Optional[int]
    title: str
    url: str
    site: Optional[str]
    category: Optional[str]
    address: Optional[str]
    guest_capacity: Optional[int]
    preview_path: Optional[str]
    is_own: bool
    description: Optional[str]
    key_facts: List[str]
    amenities: Dict[str, List[str]]
    pairs: List[PairResult]


def build_property_export_result(
    *,
    prop: Any,
    date_pairs: Sequence[Tuple[date, date]],
    rows: Sequence[str],
    states: Sequence[str],
    reasons: Sequence[Optional[str]],
) -> PropertyExportResult:
    """Build a typed export snapshot from the analysis state arrays."""
    pairs: List[PairResult] = []
    min_los_by_checkin = _min_los_categories_by_checkin(date_pairs, rows, states)
    for idx, (ci, co) in enumerate(date_pairs):
        row = rows[idx] if idx < len(rows) else ""
        state = states[idx] if idx < len(states) else "error"
        display_status = MIN_LOS if "[MinLOS]" in row else state
        pairs.append(
            PairResult(
                checkin=ci,
                checkout=co,
                nights=max(0, (co - ci).days),
                status=state,
                display_status=display_status,
                price=_parse_price_from_row(row) if state == PRICED else None,
                reason=reasons[idx] if idx < len(reasons) else None,
                min_los_nights=(
                    min_los_by_checkin.get(ci) if display_status == MIN_LOS else None
                ),
            )
        )

    return PropertyExportResult(
        id=getattr(prop, "id", None),
        title=str(getattr(prop, "title", "") or "Без названия"),
        url=str(getattr(prop, "url", "") or ""),
        site=getattr(prop, "site", None),
        category=getattr(prop, "category", None),
        address=getattr(prop, "address", None),
        guest_capacity=getattr(prop, "guest_capacity", None),
        preview_path=getattr(prop, "preview_path", None),
        is_own=bool(getattr(prop, "is_own", False)),
        description=getattr(prop, "description", None),
        key_facts=_decode_json_list(getattr(prop, "key_facts", None)),
        amenities=_decode_amenities(getattr(prop, "amenities", None)),
        pairs=pairs,
    )


def write_deep_analysis_xlsx(
    path: Path,
    results: Sequence[PropertyExportResult],
    date_pairs: Sequence[Tuple[date, date]],
    *,
    generated_at: Optional[datetime] = None,
) -> None:
    """Write the deep-analysis workbook.

    Imports are lazy so the application can still start and tests can run even
    before optional Excel dependencies are installed.
    """
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Цены"
    detail_ws = wb.create_sheet("Детализация")

    checkins = _ordered_checkins(date_pairs)
    headers = ["Объект", "Фото", "Описание", "Перечень услуг"]
    headers.extend(_date_header(d) for d in checkins)

    styles = _WorkbookStyles(
        header_fill=PatternFill("solid", fgColor="1F4E79"),
        object_fill=PatternFill("solid", fgColor="E3F2FD"),
        available_fill=PatternFill("solid", fgColor="C8E6C9"),
        unavailable_fill=PatternFill("solid", fgColor="F5F5F5"),
        minlos2_fill=PatternFill("solid", fgColor="F8C8DC"),
        minlos3_fill=PatternFill("solid", fgColor="BBDEFB"),
        warning_fill=PatternFill("solid", fgColor="FFE0B2"),
        header_font=Font(color="FFFFFF", bold=True),
        object_font=Font(color="0D2636", bold=True),
        body_font=Font(color="202020"),
        muted_font=Font(color="666666"),
        border=Border(
            left=Side(style="thin", color="D9E2EC"),
            right=Side(style="thin", color="D9E2EC"),
            top=Side(style="thin", color="D9E2EC"),
            bottom=Side(style="thin", color="D9E2EC"),
        ),
    )

    _write_matrix_sheet(ws, results, checkins, headers, styles, XLImage, Comment)
    _write_detail_sheet(detail_ws, results, styles)
    _write_legend(ws, len(results) + 3, styles)
    _finalize_matrix_sheet(ws, len(results), len(headers))
    _finalize_detail_sheet(
        detail_ws,
        sum(len(r.pairs) for r in results),
        len(_DETAIL_COLUMNS),
    )

    props = wb.properties
    props.creator = "Rental Analyzer"
    props.title = "Глубокий анализ цен"
    props.subject = "Матрица цен и детализация по датам"
    props.created = generated_at or datetime.now()

    wb.save(path)


def _write_matrix_sheet(
    ws: Any,
    results: Sequence[PropertyExportResult],
    checkins: Sequence[date],
    headers: Sequence[str],
    styles: "_WorkbookStyles",
    image_cls: Any,
    comment_cls: Any,
) -> None:
    from openpyxl.styles import Alignment

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(1, col_idx, header)
        cell.fill = styles.header_fill
        cell.font = styles.header_font
        cell.border = styles.border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, result in enumerate(results, start=2):
        ws.row_dimensions[row_idx].height = 118

        object_cell = ws.cell(row_idx, 1, _clean_cell_text(result.title))
        object_cell.fill = styles.object_fill
        object_cell.font = styles.object_font
        object_cell.border = styles.border
        object_cell.alignment = Alignment(vertical="top", wrap_text=True)
        if result.url:
            object_cell.hyperlink = result.url
            object_cell.comment = comment_cls(_clean_cell_text(result.url), "Rental Analyzer")

        photo_cell = ws.cell(row_idx, 2, "")
        photo_cell.fill = styles.object_fill
        photo_cell.border = styles.border
        photo_cell.alignment = Alignment(horizontal="center", vertical="center")
        _add_preview(ws, image_cls, result.preview_path, row_idx)

        desc_cell = ws.cell(row_idx, 3, _build_description_text(result))
        desc_cell.fill = styles.object_fill
        desc_cell.font = styles.body_font
        desc_cell.border = styles.border
        desc_cell.alignment = Alignment(vertical="top", wrap_text=True)

        amenities_cell = ws.cell(row_idx, 4, _build_amenities_text(result.amenities))
        amenities_cell.fill = styles.object_fill
        amenities_cell.font = styles.body_font
        amenities_cell.border = styles.border
        amenities_cell.alignment = Alignment(vertical="top", wrap_text=True)

        matrix = _matrix_cells_by_checkin(result.pairs)
        for offset, checkin in enumerate(checkins, start=5):
            view = matrix.get(checkin)
            cell = ws.cell(row_idx, offset)
            cell.border = styles.border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            _apply_matrix_cell(cell, view, styles)


def _write_detail_sheet(
    ws: Any,
    results: Sequence[PropertyExportResult],
    styles: "_WorkbookStyles",
) -> None:
    from openpyxl.styles import Alignment

    for col_idx, header in enumerate(_DETAIL_COLUMNS, start=1):
        cell = ws.cell(1, col_idx, header)
        cell.fill = styles.header_fill
        cell.font = styles.header_font
        cell.border = styles.border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_idx = 2
    for result in results:
        for pair in result.pairs:
            values = [
                result.title,
                result.id,
                result.site,
                result.category,
                result.url,
                pair.checkin,
                pair.checkout,
                pair.nights,
                pair.price,
                _pair_status_label(pair),
                pair.reason,
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row_idx, col_idx, value)
                cell.border = styles.border
                cell.font = styles.body_font
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if col_idx in (6, 7):
                    cell.number_format = "dd.mm.yyyy"
                elif col_idx == 9 and isinstance(value, (int, float)):
                    cell.number_format = "#,##0"
            ws.row_dimensions[row_idx].height = 42
            row_idx += 1


def _apply_matrix_cell(cell: Any, view: Optional[Dict[str, Any]], styles: "_WorkbookStyles") -> None:
    if not view:
        cell.value = "—"
        cell.fill = styles.unavailable_fill
        cell.font = styles.muted_font
        return

    price = view.get("price")
    nights = int(view.get("nights") or 0)
    status = view.get("status")

    if status == PRICED and price:
        cell.value = round(float(price))
        cell.number_format = "#,##0"
        cell.font = styles.body_font
        _apply_min_los_fill(cell, nights, styles)
        cell.comment = None if nights <= 1 else _comment_for_minlos(nights)
        return

    if status == MIN_LOS:
        cell.value = _min_los_label(nights)
        cell.font = styles.body_font
        _apply_min_los_fill(cell, nights, styles)
        cell.comment = _comment_for_minlos(nights) if nights > 1 else None
        return

    if status in {"blocked", "captcha", "network", "error", "cancelled"}:
        cell.value = _status_label(status)
        cell.fill = styles.warning_fill
        cell.font = styles.muted_font
        return

    cell.value = "—"
    cell.fill = styles.unavailable_fill
    cell.font = styles.muted_font


def _comment_for_minlos(nights: int) -> Any:
    from openpyxl.comments import Comment

    label = "2 ночи" if nights == 2 else f"{nights} ночей"
    return Comment(f"Минимальный доступный срок: {label}", "Rental Analyzer")


def _apply_min_los_fill(cell: Any, nights: int, styles: "_WorkbookStyles") -> None:
    if nights <= 1:
        cell.fill = styles.available_fill
    elif nights == 2:
        cell.fill = styles.minlos2_fill
    else:
        cell.fill = styles.minlos3_fill


def _min_los_label(nights: Optional[int]) -> str:
    if not nights or nights <= 1:
        return "MinLOS 1"
    if nights == 2:
        return "MinLOS 2"
    return "MinLOS 3+"


def _pair_status_label(pair: PairResult) -> str:
    if pair.display_status == MIN_LOS:
        return _min_los_label(pair.min_los_nights or pair.nights)
    if pair.display_status == PRICED:
        return _min_los_label(pair.nights)
    return _status_label(pair.display_status)


def _matrix_cells_by_checkin(pairs: Sequence[PairResult]) -> Dict[date, Dict[str, Any]]:
    by_checkin: Dict[date, List[PairResult]] = {}
    for pair in pairs:
        by_checkin.setdefault(pair.checkin, []).append(pair)

    matrix: Dict[date, Dict[str, Any]] = {}
    for checkin, items in by_checkin.items():
        priced = sorted(
            (p for p in items if p.status == PRICED and p.price and p.price > 0),
            key=lambda p: (p.nights, p.price or 0),
        )
        if priced:
            chosen = priced[0]
            matrix[checkin] = {
                "status": PRICED,
                "price": chosen.price,
                "nights": chosen.nights,
            }
            continue

        min_los = sorted(
            (p for p in items if p.display_status == MIN_LOS),
            key=lambda p: (p.min_los_nights or p.nights or 0, p.nights),
        )
        if min_los:
            chosen = min_los[0]
            matrix[checkin] = {
                "status": MIN_LOS,
                "price": None,
                "nights": chosen.min_los_nights or chosen.nights,
            }
            continue

        terminal = sorted(items, key=lambda p: p.nights)[0] if items else None
        if terminal:
            matrix[checkin] = {
                "status": terminal.display_status,
                "price": terminal.price,
                "nights": terminal.nights,
            }
    return matrix


def _finalize_matrix_sheet(
    ws: Any,
    result_count: int,
    column_count: int,
) -> None:
    from openpyxl.utils import get_column_letter

    ws.freeze_panes = "E2"
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = f"A1:{get_column_letter(column_count)}{max(1, result_count + 1)}"

    widths = {1: 38, 2: 18, 3: 64, 4: 54}
    for col_idx in range(1, column_count + 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = widths.get(col_idx, 13)


def _finalize_detail_sheet(
    ws: Any,
    row_count: int,
    column_count: int,
) -> None:
    from openpyxl.utils import get_column_letter

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = f"A1:{get_column_letter(column_count)}{max(1, row_count + 1)}"
    widths = [40, 10, 16, 26, 54, 16, 16, 10, 14, 18, 56]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _write_legend(ws: Any, start_row: int, styles: "_WorkbookStyles") -> None:
    from openpyxl.styles import Alignment

    legend = [
        ("Зеленый", "Доступна от 1 ночи", styles.available_fill),
        ("Розовый", "MinLOS = 2 ночи", styles.minlos2_fill),
        ("Голубой", "MinLOS = 3 и более ночей", styles.minlos3_fill),
        ("Серый", "Нет данных / недоступно", styles.unavailable_fill),
        ("Желтый", "Ошибка сети / блокировка / отмена", styles.warning_fill),
    ]
    for row_offset, (name, meaning, fill) in enumerate(legend):
        row = start_row + row_offset
        marker = ws.cell(row, 3, name)
        marker.fill = fill
        marker.border = styles.border
        marker.font = styles.body_font
        marker.alignment = Alignment(horizontal="center", vertical="center")

        text = ws.cell(row, 4, meaning)
        text.border = styles.border
        text.font = styles.body_font
        text.alignment = Alignment(vertical="center", wrap_text=True)


def _add_preview(ws: Any, image_cls: Any, preview_path: Optional[str], row_idx: int) -> None:
    if not preview_path:
        return
    path = Path(preview_path)
    if not path.is_absolute():
        path = Path.cwd() / path
    if not path.exists() or not path.is_file():
        return
    try:
        img = image_cls(str(path))
        max_width, max_height = 105, 78
        if img.width and img.height:
            scale = min(max_width / img.width, max_height / img.height, 1.0)
            img.width = int(img.width * scale)
            img.height = int(img.height * scale)
        img.anchor = f"B{row_idx}"
        ws.add_image(img)
    except Exception as exc:
        logger.debug(f"DeepAnalysis export: skipped preview {path}: {exc}")


def _build_description_text(result: PropertyExportResult) -> str:
    parts: List[str] = []
    meta = []
    if result.category:
        meta.append(str(result.category))
    if result.guest_capacity:
        meta.append(f"до {result.guest_capacity} гостей")
    if result.site:
        meta.append(str(result.site))
    if meta:
        parts.append(" · ".join(meta))
    if result.address:
        parts.append(str(result.address))
    if result.key_facts:
        parts.append(" | ".join(result.key_facts))
    if result.description:
        parts.append(str(result.description))
    return _clean_cell_text("\n".join(parts))


def _build_amenities_text(amenities: Dict[str, List[str]]) -> str:
    if not amenities:
        return ""
    lines = []
    for group, items in amenities.items():
        clean_items = [str(item).strip() for item in items if str(item).strip()]
        if clean_items:
            lines.append(f"{group}: {', '.join(clean_items)}")
    return _clean_cell_text("\n".join(lines))


def _decode_amenities(raw: Optional[str]) -> Dict[str, List[str]]:
    data = _json_loads(raw, default={})
    if not isinstance(data, dict):
        return {}
    result: Dict[str, List[str]] = {}
    for group, items in data.items():
        if not isinstance(group, str) or not isinstance(items, list):
            continue
        clean = [str(item).strip() for item in items if str(item).strip()]
        if clean:
            result[group] = clean
    return result


def _decode_json_list(raw: Optional[str]) -> List[str]:
    data = _json_loads(raw, default=[])
    if not isinstance(data, list):
        return []
    return [str(item).strip() for item in data if str(item).strip()]


def _json_loads(raw: Optional[str], *, default: Any) -> Any:
    if not raw:
        return default
    try:
        return json.loads(raw)
    except Exception:
        return default


def _parse_price_from_row(row: str) -> Optional[float]:
    if not row:
        return None
    tail = row.rsplit(";", 1)[-1].strip()
    if tail.startswith("["):
        return None
    match = _PRICE_RE.search(tail)
    if not match:
        return None
    digits = re.sub(r"\D", "", match.group(1))
    return float(digits) if digits else None


def _min_los_categories_by_checkin(
    date_pairs: Sequence[Tuple[date, date]],
    rows: Sequence[str],
    states: Sequence[str],
) -> Dict[date, int]:
    grouped: Dict[date, Dict[str, List[int]]] = {}
    for idx, (ci, co) in enumerate(date_pairs):
        nights = max(0, (co - ci).days)
        if nights <= 0:
            continue
        row = rows[idx] if idx < len(rows) else ""
        state = states[idx] if idx < len(states) else ""
        bucket = grouped.setdefault(ci, {"priced": [], "marked": []})
        if state == PRICED and _parse_price_from_row(row):
            bucket["priced"].append(nights)
        if "[MinLOS]" in row:
            bucket["marked"].append(nights)

    categories: Dict[date, int] = {}
    for ci, bucket in grouped.items():
        marked = bucket["marked"]
        if not marked:
            continue
        priced_after_mark = [
            n for n in bucket["priced"]
            if n > max(marked)
        ]
        if priced_after_mark:
            categories[ci] = min(priced_after_mark)
        else:
            categories[ci] = max(marked) + 1
    return categories


def _ordered_checkins(date_pairs: Sequence[Tuple[date, date]]) -> List[date]:
    return sorted({ci for ci, _co in date_pairs})


def _date_header(d: date) -> str:
    return d.strftime("%d.%m")


def _status_label(status: Optional[str]) -> str:
    labels = {
        PRICED: "Доступно",
        SOLD_OUT: "Недоступно",
        MIN_LOS: "MinLOS",
        "blocked": "Блокировка",
        "captcha": "Captcha",
        "network": "Сеть",
        "error": "Ошибка",
        "cancelled": "Отменено",
        "pending": "Ожидание",
        "fallback": "Fallback",
    }
    return labels.get(status or "", status or "")


def _clean_cell_text(value: Any, limit: int = 32760) -> str:
    text = "" if value is None else str(value)
    text = re.sub(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]", " ", text)
    text = text.strip()
    return text[:limit]


@dataclass(frozen=True)
class _WorkbookStyles:
    header_fill: Any
    object_fill: Any
    available_fill: Any
    unavailable_fill: Any
    minlos2_fill: Any
    minlos3_fill: Any
    warning_fill: Any
    header_font: Any
    object_font: Any
    body_font: Any
    muted_font: Any
    border: Any
