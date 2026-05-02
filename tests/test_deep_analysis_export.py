from datetime import date, timedelta
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook

from app.backend import deep_analysis as da
from app.backend.deep_analysis_export import (
    build_property_export_result,
    _parse_price_from_row,
    write_deep_analysis_xlsx,
)


def test_deep_analysis_xlsx_contains_matrix_filters_and_details():
    start = date(2026, 4, 30)
    date_pairs = da.generate_date_pairs(start, window=5)

    rows = []
    states = []
    reasons = []
    for ci, co in date_pairs:
        nights = (co - ci).days
        if ci == start and nights == 1:
            status = da._ROW_PRICED
            price = 4200
        elif ci > start and nights == 2:
            status = da._ROW_PRICED
            price = 7600
        else:
            status = da._ROW_SOLD_OUT
            price = None
        states.append(status)
        rows.append(da._format_row("Test object", ci, co, status=status, price=price))
        reasons.append("test")

    prop = SimpleNamespace(
        id=10,
        title="Test object",
        url="https://example.com/object",
        site="ostrovok",
        category="Apartment",
        address="Test address",
        guest_capacity=2,
        preview_path=None,
        is_own=False,
        description="Nice place",
        key_facts='["35 sqm", "2 guests"]',
        amenities='{"Amenities": ["Wi-Fi", "Parking"]}',
    )

    result = build_property_export_result(
        prop=prop,
        date_pairs=date_pairs,
        rows=rows,
        states=states,
        reasons=reasons,
    )
    path = Path("data") / "test_deep_analysis_export.xlsx"
    if path.exists():
        path.unlink()

    try:
        write_deep_analysis_xlsx(path, [result], date_pairs)

        wb = load_workbook(path)
        assert wb.sheetnames == ["Цены", "Детализация"]

        matrix = wb["Цены"]
        assert matrix.auto_filter.ref == "A1:H2"
        assert matrix.freeze_panes == "E2"
        assert [matrix.cell(1, col).value for col in range(1, 6)] == [
            "Объект",
            "Фото",
            "Описание",
            "Перечень услуг",
            "30.04",
        ]
        assert matrix["A2"].value == "Test object"
        assert matrix["E2"].value == 4200
        assert matrix["F2"].value == 7600

        details = wb["Детализация"]
        assert details.auto_filter.ref == "A1:K11"
        assert details["A2"].value == "Test object"
        assert details["I2"].value in (4200, 4200.0, None)
    finally:
        if path.exists():
            path.unlink()


def test_deep_analysis_export_keeps_prices_when_title_has_brackets():
    row = da._format_row("Apartment [VIP]", date(2026, 5, 1), date(2026, 5, 2),
                         status=da._ROW_PRICED, price=9300)

    assert _parse_price_from_row(row) == 9300


def test_deep_analysis_xlsx_marks_minlos_categories_and_row_sizes():
    start = date(2026, 5, 1)
    other = start + timedelta(days=1)
    date_pairs = [
        (start, start + timedelta(days=1)),
        (start, start + timedelta(days=2)),
        (start, start + timedelta(days=3)),
        (other, other + timedelta(days=1)),
        (other, other + timedelta(days=2)),
        (other, other + timedelta(days=3)),
    ]
    states = [
        da._ROW_SOLD_OUT,
        da._ROW_SOLD_OUT,
        da._ROW_PRICED,
        da._ROW_SOLD_OUT,
        da._ROW_SOLD_OUT,
        da._ROW_SOLD_OUT,
    ]
    rows = [
        da._format_row("MinLOS object", *date_pairs[0], status=states[0]),
        da._format_row("MinLOS object", *date_pairs[1], status=states[1]),
        da._format_row("MinLOS object", *date_pairs[2], status=states[2], price=18000),
        da._format_row("MinLOS object", *date_pairs[3], status=states[3]),
        da._format_row("MinLOS object", *date_pairs[4], status=states[4]),
        da._format_row("MinLOS object", *date_pairs[5], status=states[5]),
    ]
    reasons = ["test"] * len(rows)

    da._apply_minlos_marker(
        out=rows,
        states=states,
        title="MinLOS object",
        date_pairs=date_pairs,
    )

    prop = SimpleNamespace(
        id=20,
        title="MinLOS object",
        url="https://example.com/minlos",
        site="ostrovok",
        category="Apartment",
        address="Address",
        guest_capacity=2,
        preview_path=None,
        is_own=False,
        description="Long description " * 12,
        key_facts="[]",
        amenities='{"Amenities": ["Wi-Fi"]}',
    )
    result = build_property_export_result(
        prop=prop,
        date_pairs=date_pairs,
        rows=rows,
        states=states,
        reasons=reasons,
    )
    path = Path("data") / "test_deep_analysis_minlos.xlsx"
    if path.exists():
        path.unlink()

    try:
        write_deep_analysis_xlsx(path, [result], date_pairs)
        wb = load_workbook(path)
        matrix = wb[wb.sheetnames[0]]
        details = wb[wb.sheetnames[1]]

        assert matrix["E2"].value == 18000
        assert matrix["F2"].value == "MinLOS 3+"
        assert matrix.row_dimensions[2].height == 118
        assert matrix.column_dimensions["C"].width == 64
        assert details["J2"].value == "MinLOS 3+"
        assert details.row_dimensions[2].height == 42
    finally:
        if path.exists():
            path.unlink()
